import tkinter as tk
from tkinter import filedialog

import numpy as np
import openpyxl

composicoes = dict()


class Composicao:
    def __init__(self, nome, prod, equipamento, maodeobra, material, atividadeaux, tempofixo, momentodetransporte, fic):
        self.nome = nome
        self.prod = prod
        self.equipamento = equipamento
        self.maodeobra = maodeobra
        self.material = material
        self.atividadeaux = atividadeaux
        self.tempofixo = tempofixo
        self.momentodetransporte = momentodetransporte
        self.fic = int(fic)


def soma_comp(comp_code):
    fic_value = composicoes[comp_code].fic
    custo_unit = (soma_equips(comp_code) + soma_maodeobra(comp_code)) / composicoes[comp_code].prod[0]
    soma = round(
        custo_unit + custo_unit * fic_value + soma_materiais(comp_code) + soma_ativaux(comp_code) + soma_tempofixo(comp_code),
        2)  # + SomaTransporte(CompCode)

    return soma


def plan_equip_import():
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps['Sheet1']

    equips = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value is None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        data7 = sheet['H' + str(row)].value
        data8 = sheet['I' + str(row)].value
        data9 = sheet['J' + str(row)].value
        data10 = sheet['K' + str(row)].value
        equips[code] = (data1, data2, data3, data4, data5, data6, data7, data8, data9, data10)
        row = row + 1

    ps.close()

    return equips


def plan_maodeobra_import():
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Mão de Obra',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    maodeobra = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value is None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        maodeobra[code] = (data1, data2, data3, data4, data5, data6)
        row = row + 1

    ps.close()
    return maodeobra


def plan_material_import():
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Materiais',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    materiais = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value is None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        if sheet['D' + str(row)].value == '-':
            data3 = 0
        else:
            data3 = sheet['D' + str(row)].value
        materiais[code] = (data1, data2, data3)
        row = row + 1

    ps.close()
    return materiais


def plan_precos_import():
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Composições',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    preco = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value is None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data3 = sheet['D' + str(row)].value
        preco[code] = data3
        row = row + 1

    ps.close()
    return preco


def soma_equips(comp_code):
    soma = 0

    comp_equips = composicoes[comp_code].equipamento
    if len(comp_equips) != 0:
        for each in comp_equips:
            equip_code = each[0]
            equip_quant = each[1]
            equip_prod = each[2]
            equip_improd = each[3]
            soma = soma + round(equip_quant * (equip_prod * equips[equip_code][8] + equip_improd * equips[equip_code][9]), 4)

    return round(soma, 4)


def soma_maodeobra(comp_code):
    soma = 0
    comp_maodeobras = composicoes[comp_code].maodeobra
    if len(comp_maodeobras) != 0:
        for each in comp_maodeobras:
            maodeobra_code = each[0]
            maodeobra_quant = each[1]
            soma = soma + round(maodeobra_quant * maodeobra[maodeobra_code][4], 4)

    return round(soma, 4)


def soma_materiais(comp_code):
    soma = 0
    comp_materiais = composicoes[comp_code].material
    if len(comp_materiais) != 0:
        for each in comp_materiais:
            material_code = each[0]
            material_quant = each[1]
            soma = soma + round(material_quant * materiais[material_code][2], 4)

    return round(soma, 4)


def soma_ativaux(comp_code):
    soma = 0
    comp_ativaux = composicoes[comp_code].atividadeaux
    if len(comp_ativaux) != 0:
        for each in comp_ativaux:
            ativaux_code = each[0]
            ativaux_quant = each[1]
            soma = soma + round(ativaux_quant * soma_comp(ativaux_code), 4)

    return round(soma, 4)


def soma_tempofixo(comp_code):
    soma = 0
    comp_tempofixo = composicoes[comp_code].tempofixo
    if len(comp_tempofixo) != 0:
        for each in comp_tempofixo:
            tempofixo_code = each[1]
            tempofixo_quant = each[2]
            soma = soma + round(tempofixo_quant * soma_comp(tempofixo_code), 4)

    return round(soma, 4)


def soma_transporte(comp_code):
    soma = 0
    comp_transporte = composicoes[comp_code].momentodetransporte
    if len(comp_transporte) != 0:
        for each in comp_transporte:
            transporte_code = each[4]
            transporte_quant = each[1]
            soma = soma + round(transporte_quant * soma_comp(transporte_code), 4)

    return round(soma, 4)


equips = plan_equip_import()
maodeobra = plan_maodeobra_import()
materiais = plan_material_import()
preco = plan_precos_import()

list = np.load("comps5.npy", allow_pickle=True)
# loadFIC = np.load("FIC-PR.npy", allow_pickle=True)
dict = list[()]
for key in dict:
    composicoes[key] = Composicao(dict[key][0], dict[key][1], dict[key][2], dict[key][3], dict[key][4], dict[key][5],
                                  dict[key][6], dict[key][7], dict[key][8])

CompCode = '0308265'
print('Equip', soma_equips(CompCode))
print('Mão de obra', soma_maodeobra(CompCode))
print('Material', soma_materiais(CompCode))
print('Atividades Auxiliares', soma_ativaux(CompCode))
print('Tempo Fixo', soma_tempofixo(CompCode))
print('Momento de Transporte', soma_transporte(CompCode))

print(round(soma_comp(CompCode), 4))

# for code in preco:
#    if code != '0919002' and code != '0919210' and code != '7119788':
#        soma = SOMAComp(code)
#        verif = preco[code] == soma
#        if not verif:
#            print(code, preco[code], soma, verif)
#    else:
#        continue


# print(dict)
# file = open("dict.txt", "w")
# str_dictionary = repr(dict)
# file.write(str_dictionary + "\n")
# file.close()
# print(FIC)
# print(materiais)
# print(equips)
# print(maodeobra)
# print(preco)

# print(soma - SomaTransporte(CompCode))
