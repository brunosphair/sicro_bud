import openpyxl
import json
import tkinter as tk
import inspect
from tkinter import filedialog

'''
This module is destined to import the data of analytical compositions report
from the *.xlsx sicro file to a *.json file to be used in the sicro.py module.
This *.json file contains all the information necessary to calculate any
composition of SICRO.
'''

class SicroWorksheet:
    def __init__(self):
        self.comps = {}

    def sicro_xlsx_import(self, filepath=None):
        '''
        Selection and load of the *.xlsx analytical compositions report
        '''
        if not filepath:
            self.filepath = filedialog.askopenfilename(
            title='Selecione o arquivo referente ao Relatório Analítico de Composições de Custos',
            filetypes=[("Excel files", "*.xlsx")])
        else:
            self.filepath = filepath
        root = tk.Tk()
        root.withdraw()
        print("Loading Workbook...")
        ps = openpyxl.load_workbook(self.filepath)
        self.sheet = ps.active
        self.first_row_mapping()
        self.wb_data_storaging()
        ps.close()
        with open('comps2.json', 'w') as fp:
            json.dump(self.comps, fp)
        print("Done!")

    def first_row_mapping(self):
        '''
        Mapping all the first rows of the compositions in the Excel file.
        '''
        target_rows = []
        for row in self.sheet.iter_rows(min_row=self.sheet.min_row,
                                        max_row=self.sheet.max_row,
                                        max_col=1):
            for cell in row:
                if cell.value == "CGCIT":
                    target_rows.append(cell.row)

        self.target_rows = target_rows

    def wb_data_storaging(self):
        '''
        Search of the necessary data in the Excel file and storing in a
        dictionary.
        '''
        for row in self.target_rows:
            comp_number, comp_list = self.get_comp_data(row)
            if comp_number:
                self.comps[comp_number] = comp_list
                #print(row, " ", row * 100 / self.sheet.max_row, "%")

    def get_comp_data(self, row):
        '''
        Get all the data of the specified composition, what initiates in the
        specified row
        '''

        fic = self.get_fic(row)
        
        # Find and store the composition data: number and name
        comp_number = self.sheet['A' + str(row + 3)].value
        comp_name = self.sheet['B' + str(row + 3)].value
        try:
            prod = self.get_prod(row)
            row += 6
            equip, row = self.get_equip(row)
            row += 2
            labor, row = self.get_labor(row)
            row += 6
            material, row = self.get_material(row)
            row += 2
            aux_activity, row = self.get_aux_activity(row)
            row += 3
            fixed_time, row = self.get_fixed_time(row)
            row += 3
            transport, row = self.get_transport(row)
        except AssertionError:
            print("Oops! The composition " + comp_number + " was not imported"
                  " because of an error in ." + inspect.trace()[-1][3] + 
                  " function. Bad worksheet formatting.")
            return None, None

                    # Add all the data in a dictionary, where the key is the composition number
        return comp_number,(comp_name,
                            prod,
                            equip,
                            labor,
                            material,
                            aux_activity,
                            fixed_time,
                            transport,
                            fic)

    def get_fic(self, row):
        '''
        Gets the FIC in the specified row.
        '''
        
        if self.sheet['G' + str(row)].value == "FIC":
            fic = self.sheet['H' + str(row)].value
        else:
            fic = 0
        
        return fic

    def get_prod(self, row):
        '''
        Find and store the composition productivity data
        '''
        text = self.sheet['G' + str(row + 2)].value
        if text:
            text = text.strip()
        assert text == 'Produção da equipe'

        prod = self.sheet['H' + str(row + 2)].value
        prod_unit = self.sheet['I' + str(row + 2)].value
        return (prod, prod_unit)

    def get_equip(self, row):
        '''
        Find and store the data relating to the composition equipments.
        '''
        text = self.sheet['A' + str(row - 2)].value
        if text:
            text = text.strip()
        assert text == 'A - EQUIPAMENTOS'

        equipment = []
        while self.sheet['A' + str(row)].value is not None:
            equipment_code = self.sheet['A' + str(row)].value
            equipment_qtt = self.sheet['C' + str(row)].value
            equip_prod_time = self.sheet['D' + str(row)].value
            equip_improd_time = self.sheet['E' + str(row)].value
            equipment.append((equipment_code, equipment_qtt, equip_prod_time, equip_improd_time))
            row += 1

        return equipment, row

    def get_labor(self, row):
        '''
        Find and store in a list the data relating to the composition labor.
        '''
        text = self.sheet['A' + str(row - 1)].value
        if text:
            text = text.strip()
        assert text == 'B - MÃO DE OBRA'
        
        labor = []
        while self.sheet['A' + str(row)].value is not None:
            labor_code = self.sheet['A' + str(row)].value
            labor_qtt = self.sheet['C' + str(row)].value
            labor.append((labor_code, labor_qtt))
            row += 1
        
        return labor, row

    def get_material(self, row):
        '''
        Find and store in a list the data relating to the composition materials.
        '''
        text = self.sheet['A' + str(row - 1)].value
        if text:
            text = text.strip()
        assert text == 'C - MATERIAL'

        material = []
        while self.sheet['A' + str(row)].value is not None:
            material_code = self.sheet['A' + str(row)].value
            material_qtt = self.sheet['C' + str(row)].value
            material.append((material_code, material_qtt))
            row = row + 1
        
        return material, row

    def get_aux_activity(self, row):
        '''
        Find and store in a list the data relating to the composition auxiliary
        activities.
        '''
        text = self.sheet['A' + str(row - 1)].value
        if text:
            text = text.strip()
        assert text == 'D - ATIVIDADES AUXILIARES'

        aux_activity = []
        while self.sheet['A' + str(row)].value is not None:
            aux_activity_code = self.sheet['A' + str(row)].value
            aux_activity_qtt = self.sheet['C' + str(row)].value
            aux_activity.append((aux_activity_code, aux_activity_qtt))
            row += 1

        return aux_activity, row

    def get_fixed_time(self, row):
        '''
        Find and store the composition fixed time data
        '''
        text = self.sheet['A' + str(row - 1)].value
        if text:
            text = text.strip()
        assert text == 'E - TEMPO FIXO'

        fixed_time = []
        while self.sheet['A' + str(row)].value is not None:
            fixed_time_code = self.sheet['A' + str(row)].value
            fixed_time_material_code = self.sheet['C' + str(row)].value
            fixed_time_qtt = self.sheet['D' + str(row)].value
            fixed_time.append((fixed_time_code, fixed_time_material_code, fixed_time_qtt))
            row = row + 1
        
        return fixed_time, row

    def get_transport(self, row):
        '''
        Find and store the composition transport data.
        '''
        text = self.sheet['A' + str(row - 2)].value
        if text:
            text = text.strip()        
        assert text == 'F - MOMENTO DE TRANSPORTE'

        transport = []
        while self.sheet['A' + str(row)].value is not None:
            transport_code = self.sheet['A' + str(row)].value
            transport_qtt = self.sheet['C' + str(row)].value
            transport_dirt_road = self.sheet['E' + str(row)].value
            transport_stone_road = self.sheet['F' + str(row)].value
            transport_asphalt = self.sheet['G' + str(row)].value
            transport.append((transport_code, transport_qtt, transport_dirt_road,
                                transport_stone_road, transport_asphalt))
            row = row + 1

        return transport, row

if __name__ == '__main__':
    ws = SicroWorksheet()
    ws.sicro_xlsx_import()
