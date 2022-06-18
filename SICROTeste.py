import numpy as np
import openpyxl
import tkinter as tk
from tkinter import filedialog
import numpy as np
import pandas as pd


def SOMAComp(CompCode):

    if FIC[CompCode] == None:
        FICValue = 0
    else:
        FICValue = FIC[CompCode]

    CustoUnit = (SomaEquips(CompCode) + SomaMaodeObra(CompCode)) / dict[CompCode][1][0]
    soma = round(CustoUnit + CustoUnit * FICValue + SomaMateriais(CompCode) + SomaAtivAux(CompCode) + SomaTempoFixo(CompCode), 2) # + SomaTransporte(CompCode)

    return soma


def PLANEquipImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps['Sheet1']

    equips = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        data7 = sheet['H' + str(row)].value
        data8 = sheet['I' + str(row)].value
        data9 = sheet['J' + str(row)].value
        data10 = sheet['K' + str(row)].value
        equips[code] = (data1, data2, data3, data4, data5, data6, data7, data8, data9, data10)
        row = row + 1

    ps.close()

    return equips

def PLANMaodeObraImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Mão de Obra',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    maodeobra = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        maodeobra[code] = (data1, data2, data3, data4, data5, data6)
        row = row + 1

    ps.close()
    return maodeobra

def PLANMaterialImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Materiais',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    materiais = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        if sheet['D' + str(row)].value == '-':
            data3 = 0
        else:
            data3 = sheet['D' + str(row)].value
        materiais[code] = (data1, data2, data3)
        row = row + 1

    ps.close()
    return materiais

def PLANPrecosImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Composições',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    preco = {}

    row = 1

    while sheet['A' + str(row)].value != "Código":
        row = row + 1

    row = row + 1

    while sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data3 = sheet['D' + str(row)].value
        preco[code] = data3
        row = row + 1

    ps.close()
    return preco


def SomaEquips(CompCode):

    soma = 0

    COMPEquips = dict[CompCode][2]
    if len(COMPEquips) != 0:
        for each in COMPEquips:
            EQUIPCode = each[0]
            EQUIPQuant = each[1]
            EQUIPProd = each[2]
            EQUIPImprod = each[3]
            soma = soma + round(EQUIPQuant * (EQUIPProd * equips[EQUIPCode][8] + EQUIPImprod * equips[EQUIPCode][9]), 4)

    return round(soma, 4)

def SomaMaodeObra(CompCode):

    soma = 0
    COMPMaodeObras = dict[CompCode][3]
    if len(COMPMaodeObras) != 0:
        for each in COMPMaodeObras:
            MAODEOBRACode = each[0]
            MAODEOBRAQuant = each[1]
            soma = soma + round(MAODEOBRAQuant * maodeobra[MAODEOBRACode][4], 4)

    return round(soma, 4)

def SomaMateriais(CompCode):

    soma = 0
    COMPMateriais = dict[CompCode][4]
    if len(COMPMateriais) != 0:
        for each in COMPMateriais:
            MATERIALCode = each[0]
            MATERIALQuant = each[1]
            soma = soma + round(MATERIALQuant * materiais[MATERIALCode][2], 4)

    return round(soma, 4)

def SomaAtivAux(CompCode):

    soma = 0
    COMPAtivAux = dict[CompCode][5]
    if len(COMPAtivAux) != 0:
        for each in COMPAtivAux:
            ATIVAUXCode = each[0]
            ATIVAUXQuant = each[1]
            soma = soma + round(ATIVAUXQuant * SOMAComp(ATIVAUXCode), 4)

    return round(soma, 4)

def SomaTempoFixo(CompCode):

    soma = 0
    COMPTempoFixo = dict[CompCode][6]
    if len(COMPTempoFixo) != 0:
        for each in COMPTempoFixo:
            TEMPOFIXOCode = each[1]
            TEMPOFIXOQuant = each[2]
            soma = soma + round(TEMPOFIXOQuant * SOMAComp(TEMPOFIXOCode), 4)

    return round(soma, 4)

def SomaTransporte(CompCode):

    soma = 0
    COMPTransporte = dict[CompCode][7]
    if len(COMPTransporte) != 0:
        for each in COMPTransporte:
            TRANSPORTECode = each[4]
            TRANSPORTEQuant = each[1]
            soma = soma + round(TRANSPORTEQuant * SOMAComp(TRANSPORTECode), 4)

    return round(soma, 4)


def GetProdutiv():

    wb = openpyxl.Workbook()
    ws = wb.active

    composicoes = [['5502985', 79764], ['5502985', 99444], ['5502978', 69519], ['4011492', 1415.91], ['4011214', 923.42], ['4011209', 6156.13], ['4011214', 1999.18], ['4011228', 3065.41], ['4011209', 13327.85], ['4011228', 1929.33], ['4011209', 12058.3], ['4011228', 309.83], ['4011209', 2213.04], ['4011214', 1040.63], ['4011228', 1769.07], ['4011209', 10406.27], ['4011214', 319.33], ['4011228', 574.79], ['4011209', 3193.25], ['1107892', 510.45], ['2003842', 723.096], ['4805750', 485.55], ['3103302', 581.166], ['4805755', 60.2784], ['1107892', 25.488], ['2003842', 36.1152], ['2004521', 74.9664], ['3108018', 200.88], ['0407820', 523.6], ['1107895', 16], ['3103302', 200], ['0407820', 2225.3], ['1107895', 63.58], ['3103302', 809.2], ['0407820', 184.8], ['1107895', 5.28], ['3103302', 67.2], ['4805755', 114.3255], ['1107892', 56.4925], ['2003842', 80.047], ['4413996', 651.1], ['2004522', 180.9675], ['3108022', 381.085], ['4805755', 316.7085], ['1107892', 156.4975], ['2003842', 221.749], ['4413996', 1803.7], ['2004522', 501.3225], ['3108022', 1055.695], ['4805755', 66.78], ['1107892', 32.34], ['2003842', 45.808], ['4413996', 420], ['2004522', 93.1], ['3108022', 222.6], ['4805755', 44.775], ['1107892', 22.125], ['2003842', 31.35], ['4413996', 255], ['2004522', 70.875], ['3108022', 149.25], ['2009619', 331.47], ['1109669', 5.22], ['0407819', 356.7], ['1107892', 21.75], ['1107896', 5.22], ['3103302', 269.7], ['4805755', 60], ['0407819', 144.96], ['1107892', 26.4], ['1107896', 1.104], ['4805751', 180], ['3103302', 260.16], ['4805755', 65], ['0407819', 157.04], ['1107892', 28.6], ['1107896', 1.196], ['4805751', 195], ['3103302', 281.84], ['0407820', 41], ['1107892', 14.1], ['3103302', 119.3], ['0407820', 17], ['1107892', 1.74], ['3103302', 15.05], ['4805755', 1.5], ['1107892', 13], ['2003842', 52.7245], ['4805750', 4.5], ['3103302', 13], ['4805755', 18.7], ['1107892', 19.25], ['4805750', 39.6], ['3103302', 83.6], ['1107892', 2.94], ['4805750', 4.69], ['3103302', 18.97], ['1109671', 1.122], ['1106165', 102.68], ['3103302', 340], ['1109671', 0.9339], ['1106165', 85.466], ['3103302', 283], ['1109671', 0.05775], ['1106165', 5.285], ['3103302', 17.5], ['1109671', 3.1949], ['1106165', 167.175], ['3103302', 445.8], ['1109671', 0.7095], ['1106165', 37.125], ['3103302', 99], ['1109671', 2.123], ['1106165', 118.888], ['3103302', 270.2], ['1109671', 0.396], ['1106165', 22.176], ['3103302', 50.4], ['1109671', 1.16865], ['1106165', 63.918], ['3103302', 127.2], ['1109671', 0.33075], ['1106165', 18.09], ['3103302', 36], ['4805757', 14255], ['4815671', 12821], ['3108005', 19695.99], ['0407819', 369116.89], ['1107890', 3675.23], ['2408149', 329287.3], ['2408149', 87882.9], ['5502978', 48950], ['1505860', 295], ['3108005', 53.2], ['0407819', 1190.4], ['1107890', 14.88], ['3108005', 32.48], ['0407819', 324.8], ['1107890', 3.25], ['3108005', 5.04], ['0407819', 111.72], ['1107890', 1.4], ['3108005', 13.82], ['0407819', 870.39], ['1107890', 10.88], ['3108005', 280.55], ['0407819', 4809.6], ['1107890', 60.12], ['3108005', 0], ['0407819', 2679.52], ['1107890', 33.49], ['1107890', 996.68], ['3108005', 5022.86], ['0407819', 112148], ['3108005', 3565], ['0407819', 59855.9], ['1107890', 668.92], ['3108005', 91.2], ['0407819', 1094.4], ['1107890', 13.68], ['3108005', 161.68], ['0407819', 1616.8], ['1107890', 16.17], ['3108005', 161.7], ['0407819', 1155], ['1107890', 11.55], ['3108005', 76.96], ['0407819', 436.8], ['1107890', 4.37], ['3108005', 133.37], ['0407819', 2921.18], ['1107890', 36.51], ['2408149', 773.7], ['2408149', 19249], ['3108005', 94.2], ['0407819', 1368], ['1107890', 17.1], ['2408149', 135162.95], ['2408149', 5320], ['3108005', 238.21], ['0407819', 5810], ['1107890', 58.1], ['2408149', 179717.48], ['2408149', 17971.7], ['3108005', 128.8], ['0407819', 3143], ['1107890', 31.43], ['2408149', 31091.4], ['2408149', 116212.1], ['2408149', 11621.21], ['5502978', 201212], ['1505860', 52616.4], ['1505860', 134231], ['5405983', 7350.75], ['5405984', 2106], ['5406043', 3017.25], ['5406023', 172.1], ['5406025', 144.6], ['5406027', 110.1], ['5406029', 108], ['5406031', 836.9], ['0705397', 19], ['4011209', 294]]

    ws['B2'].value = 'CÓD'
    ws['C2'].value = 'Descrição do Serviço'
    ws['D2'].value = 'Qtde.'
    ws['E2'].value = 'Qtde Aux'
    ws['F2'].value = 'Prod'
    ws['G2'].value = 'Duração (dias)'
    ws['H2'].value = 'Duração MOD'
    ws['I2'].value = 'CÓD'
    ws['J2'].value = 'EQUIPAMENTO/MÃO DE OBRA'
    ws['K2'].value = 'Qtde Unit'
    ws['L2'].value = 'Qtde Total'

    row = 3

    for comp in composicoes:
        code = comp[0]
        quant = comp[1]
        ws.cell(row=row, column=2).value = code
        descricao = dict[code][0]
        ws.cell(row=row, column=3).value = descricao
        ws.cell(row=row, column=4).value = quant
        produtiv = dict[code][1][0]
        ws.cell(row=row, column=6).value = produtiv
        ws.cell(row=row, column=7).value = '=D' + str(row) + '/F' + str(row) + '/8'
        ws.cell(row=row, column=8).value = quant / produtiv / 8

        rowcomp = row
        row = row + 1

        COMPEquips = dict[code][2]
        for each in COMPEquips:
            cod = each[0]
            qtde = each[1]
            ws.cell(row=row, column=9).value = cod
            ws.cell(row=row, column=10).value = equips[cod][0]
            ws.cell(row=row, column=11).value = qtde
            ws.cell(row=row, column=12).value = '=K' + str(row) + '*($G$' + str(rowcomp) + '/$H$' + str(rowcomp) + ')'
            row = row + 1
#            print(CompCode, produtiv, each[0], each[1])
        COMPMaodeObras = dict[code][3]
        for each in COMPMaodeObras:
            cod = each[0]
            qtde = each[1]
            ws.cell(row=row, column=9).value = cod
            ws.cell(row=row, column=10).value = maodeobra[cod][0]
            ws.cell(row=row, column=11).value = qtde
            ws.cell(row=row, column=12).value = '=K' + str(row) + '*($G$' + str(rowcomp) + '/$H$' + str(rowcomp) + ')'
            row = row + 1
#            print(CompCode, produtiv, each[0], each[1])
        COMPAux = dict[code][5]
        for each in COMPAux:
            codeaux = each[0]
            quant = each[1]
            row = GetProdutivAux(codeaux, quant, row, rowcomp, ws)

    wb.save("Produtiv.xlsx")


def GetProdutivAux(codeaux, quant, row, rowcomp, ws):

    COMPEquips = dict[codeaux][2]
    codeauxname = dict[codeaux][0]
    ws.cell(row=row, column=3).value = codeauxname
    ws.cell(row=row, column=5).value = '=' + str(quant) + '*D' + str(rowcomp)
    produtiv = dict[codeaux][1][0]
    ws.cell(row=row, column=6).value = produtiv
    ws.cell(row=row, column=7).value = '=E' + str(row) + '/F' + str(row) + '/8'
    ws.cell(row=row, column=8).value = '=H' + str(rowcomp)
    for each in COMPEquips:
        cod = each[0]
        qtde = each[1]
        ws.cell(row=row, column=9).value = cod
        ws.cell(row=row, column=10).value = equips[cod][0]
        ws.cell(row=row, column=11).value = qtde
        ws.cell(row=row, column=12).value = '=K' + str(row) + '*($G$' + str(rowcomp) + '/$H$' + str(rowcomp) + ')'
        row = row + 1
    #            print(CompCode, produtiv, each[0], each[1])
    COMPMaodeObras = dict[codeaux][3]
    for each in COMPMaodeObras:
        cod = each[0]
        qtde = each[1]
        ws.cell(row=row, column=9).value = cod
        ws.cell(row=row, column=10).value = maodeobra[cod][0]
        ws.cell(row=row, column=11).value = qtde
        ws.cell(row=row, column=12).value = '=K' + str(row) + '*($G$' + str(rowcomp) + '/$H$' + str(rowcomp) + ')'
        row = row + 1
    COMPAux = dict[codeaux][5]
    for each in COMPAux:
        codeaux = each[0]
        quant = each[1] * quant
        GetProdutivAux(codeaux, quant, row, rowcomp, ws)

    return row


equips = PLANEquipImport()
maodeobra = PLANMaodeObraImport()
materiais = PLANMaterialImport()
#preco = PLANPrecosImport()

list = np.load("comps4.npy", allow_pickle=True)
#loadFIC = np.load("FIC-PR.npy", allow_pickle=True)
dict = list[()]
#FIC = loadFIC[()]


GetProdutiv()


#CompCode = '0606845'
#print(dict[CompCode][1])
#print('Equip', SomaEquips(CompCode))
#print('Mão de obra', SomaMaodeObra(CompCode))
#print('Material', SomaMateriais(CompCode))
#print('Atividades Auxiliares', SomaAtivAux(CompCode))
#print('Tempo Fixo', SomaTempoFixo(CompCode))
#print('Momento de Transporte', SomaTransporte(CompCode))

#print(round(SOMAComp(CompCode),4))

#for code in preco:
#    if code != '0919002' and code != '0919210' and code != '7119788':
#        soma = SOMAComp(code)
#        verif = preco[code] == soma
#        if not verif:
#            print(code, preco[code], soma, verif)
#    else:
#        continue


#print(dict)
#file = open("dict.txt", "w")
#str_dictionary = repr(dict)
#file.write(str_dictionary + "\n")
#file.close()
#print(FIC)
#print(materiais)
#print(equips)
#print(maodeobra)
#print(preco)

#print(soma - SomaTransporte(CompCode))
