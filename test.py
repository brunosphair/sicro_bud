import numpy as np
import openpyxl
import tkinter as tk
from tkinter import filedialog
import numpy as np
import pandas as pd

root = tk.Tk()
root.withdraw()
filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                          filetypes=[("Excel files", "*.xlsx")])

ps = openpyxl.load_workbook(filepath)
sheet = ps.active

FIC = {}
TargetRows = []

row = 1

if sheet['A' + str(row)].value != "CGCIT":
    row = row + 1

#for row in sheet.iter_rows("A"):
#    for cell in row:
#        if cell.value == "CGCIT":
#            print(sheet.cell(row=cell.row, column=1).value)
for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, max_col=1):
    for cell in row:
        if cell.value == "CGCIT":
            TargetRows.append(cell.row)
            print(cell.row)

for row in TargetRows:
    COMPCode = sheet['A' + str(row + 3)].value
    FICValue = sheet['H' + str(row + 1)].value
    FIC[COMPCode] = FICValue
    print(row)

print(FIC)

#while row <= 1000:
#    if sheet['A' + str(row)].value == "CGCIT":
#        print(row)
#        row = row + 1
#        FICValue = sheet['H' + str(row)].value
#        row = row + 2
#        code = sheet['A' + str(row)].value
#        FIC[code] = FICValue
#        row = row + 15
#        TargetRows.append((row))
#    row = row + 1

ps.close()

print(FIC)
np.save('FIC-PR.npy', FIC)