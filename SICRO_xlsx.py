import openpyxl
import pandas as pd

ps = openpyxl.load_workbook('D:\SICRO\Planilha editavel\sicro-planilha-editavel-ref-out-2021.xlsx')
sheet = ps['Sheet1']

comps = []
nrows = sheet.max_row

for row in range(1,sheet.max_row + 1):
    if sheet['A' + str(row)].value == "CGCIT":
        comps.append(sheet['A' + str(row + 3)].value)

print(comps)

