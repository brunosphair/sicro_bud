import openpyxl
import numpy as np
import pandas as pd

ps = openpyxl.load_workbook('D:\SICRO\pr-01-2022\PR 01-2022 Relatório Analítico de Composições de Custos.xlsx')
sheet = ps['Sheet1']

comps = {}
nrows = sheet.max_row

row = 1

# sheet.max_row

TargetRows = []

for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, max_col=1):
    for cell in row:
        if cell.value == "CGCIT":
            TargetRows.append(cell.row)
            print(cell.row)

for row in TargetRows:
    if sheet['A' + str(row)].value == "CGCIT":
        if sheet['G' + str(row)].value == "FIC":
            fic = sheet['H' + str(row)].value
        else:
            fic = 0
        row = row + 3
        composicao = sheet['A' + str(row)].value
        nome = sheet['B' + str(row)].value
        producao = sheet['H' + str(row - 1)].value
        unidade = sheet['I' + str(row - 1)].value
        prod = (producao, unidade)
        equipamento = []
        row = row + 3
        while sheet['A' + str(row)].value is not None:
            equipamentocod = sheet['A' + str(row)].value
            equipamentoqtde = sheet['C' + str(row)].value
            equipamentoop = sheet['D' + str(row)].value
            equipamentoimp = sheet['E' + str(row)].value
            equipamento.append((equipamentocod, equipamentoqtde, equipamentoop, equipamentoimp))
            row = row + 1
        row = row + 2
        maodeobra = []
        while sheet['A' + str(row)].value is not None:
            maodeobracod = sheet['A' + str(row)].value
            maodeobraqtde = sheet['C' + str(row)].value
            maodeobra.append((maodeobracod, maodeobraqtde))
            row = row + 1
        row = row + 6
        material = []
        while sheet['A' + str(row)].value is not None:
            materialcod = sheet['A' + str(row)].value
            materialqtde = sheet['C' + str(row)].value
            material.append((materialcod, materialqtde))
            row = row + 1
        row = row + 2
        atividadeaux = []
        while sheet['A' + str(row)].value is not None:
            atividadeauxcod = sheet['A' + str(row)].value
            atividadeauxqtde = sheet['C' + str(row)].value
            atividadeaux.append((atividadeauxcod, atividadeauxqtde))
            row = row + 1
        row = row + 3
        tempofixo = []
        while sheet['A' + str(row)].value is not None:
            tempofixocod = sheet['A' + str(row)].value
            tempofixomat = sheet['C' + str(row)].value
            tempofixoqtde = sheet['D' + str(row)].value
            tempofixo.append((tempofixocod, tempofixomat, tempofixoqtde))
            row = row + 1
        row = row + 3
        momentodetransporte = []
        while sheet['A' + str(row)].value is not None:
            momentodetransportecod = sheet['A' + str(row)].value
            momentodetransporteqtde = sheet['C' + str(row)].value
            momentodetransporteLN = sheet['E' + str(row)].value
            momentodetransporteRP = sheet['F' + str(row)].value
            momentodetransporteP = sheet['G' + str(row)].value
            momentodetransporte.append((momentodetransportecod, momentodetransporteqtde, momentodetransporteLN,
                                        momentodetransporteRP, momentodetransporteP))
            row = row + 1
        # print("COMP:", composicao, "EQUIP:", equipamento, "MAODEOBRA:", maodeobra, "MAT:", material, "AUX:",
        #     atividadeaux, "TFIXO:", tempofixo, "MTRANSP:", momentodetransporte)

        comps[composicao] = (
            nome, prod, equipamento, maodeobra, material, atividadeaux, tempofixo, momentodetransporte, fic)

        print(row, " ", row * 100 / sheet.max_row, "%")
    row = row + 1

print(comps)
np.save('comps5.npy', comps)
print("Pronto")
