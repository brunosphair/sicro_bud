import openpyxl
import tkinter as tk
from tkinter import filedialog
import numpy as np
import pandas as pd


root = tk.Tk()
root.withdraw()
filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos', filetypes=[("Excel files", "*.xlsx")])

ps = openpyxl.load_workbook(filepath)
sheet = ps['Sheet1']

equips = {}
nrows = sheet.max_row

row = 4

#sheet.max_row

while row <= sheet.max_row:
    code = sheet['A' + str(row)].value
    data1 = sheet['B' + str(row)].value
    data2 = sheet['C' + str(row)].value
    data3 = sheet['D' + str(row)].value
    data4 = sheet['E' + str(row)].value
    data5 = sheet['F' + str(row)].value
    data6 = sheet['G' + str(row)].value
    data7 = sheet['H' + str(row)].value
    data8 = sheet['I' + str(row)].value
    data9 = sheet['J' + str(row)].value
    data10 = sheet['K' + str(row)].value
    equips[code] = (data1, data2, data3, data4, data5, data6, data7, data8, data9, data10)
    row = row + 1

print(equips)
