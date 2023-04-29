import openpyxl
import tkinter as tk
from tkinter import filedialog
import json

def get_xlsx_materials(filepath=None):
    if not filepath:
        root = tk.Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(title="Selecione o arquivo referente ao Relatório Sintético de Materiais",
                                                filetypes=[("Excel files", "*.xlsx")]) 
    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    table_rows = []
    row = 1

    while sheet['A' + str(row)].value != "Código":
        row += 1

    row += 1

    while sheet['A' + str(row)].value is None:
        row += 1

    while row <= sheet.max_row:
        materials = {}
        materials["code"] = sheet['A' + str(row)].value
        materials["description"] = sheet['B' + str(row)].value
        materials["unit"] = sheet['C' + str(row)].value
        if sheet['D' + str(row)].value == '-':
            materials["price"] = 0
        else:
            materials["price"] = sheet['D' + str(row)].value
        table_rows.append(materials)
        row = row + 1

    ps.close()

    return table_rows

def get_xlsx_equipments(filepath=None):
    '''
    Imports the equipments data from the *.xlsx file to a list of dictionaries.
    '''

    if not filepath:
        root = tk.Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                            filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    table_rows = []
    row = 1

    while sheet['A' + str(row)].value != "Código":
        row += 1
    row += 1

    while sheet['A' + str(row)].value is None:
        row += 1

    while row <= sheet.max_row:
        equips = {}
        equips["code"] = sheet['A' + str(row)].value
        equips["description"] = sheet['B' + str(row)].value
        equips["acquisition_price"] = sheet['C' + str(row)].value
        equips["depreciation"] = sheet['D' + str(row)].value
        equips["capital_opportunity"] = sheet['E' + str(row)].value
        equips["insurance_and_taxes"] = sheet['F' + str(row)].value
        equips["maintenance"] = sheet['G' + str(row)].value
        equips["operation"] = sheet['H' + str(row)].value
        equips["operation_labor"] = sheet['I' + str(row)].value
        equips["productive_cost"] = sheet['J' + str(row)].value
        equips["unproductive_cost"] = sheet['K' + str(row)].value
        table_rows.append(equips)
        row = row + 1

    ps.close()

    return table_rows

def get_xlsx_labors(filepath=None):
    '''
    Imports the labors data from the *.xlsx file to a dictionary.
    '''

    if not filepath:
        root = tk.Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Mão de Obra',
                                            filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    table_rows = []
    row = 1

    while sheet['A' + str(row)].value != "Código":
        row += 1
    row += 1

    while sheet['A' + str(row)].value is None:
        row += 1

    while row <= sheet.max_row:
        labor = {}
        labor["code"] = sheet['A' + str(row)].value
        labor["description"] = sheet['B' + str(row)].value
        labor["unit"] = sheet['C' + str(row)].value
        labor["salary"] = sheet['D' + str(row)].value
        labor["total_charges"] = sheet['E' + str(row)].value
        labor["cost"] = sheet['F' + str(row)].value
        labor["dangerousness_and_insalubrity"] = sheet['G' + str(row)].value
        table_rows.append(labor)
        row += 1

    ps.close()

    return table_rows

def get_json_gd_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_general_data = []
    for key in dict_import:
        general_data={}
        general_data["comp_code"] = key
        general_data["description"] = dict_import[key][0]
        general_data["productivity_unit"] = dict_import[key][1][1]
        general_data["productivity"] = dict_import[key][1][0]
        general_data["fic"] = round(float(dict_import[key][8]))
        sicro_general_data.append(general_data)
        
        # comps[key] = Composition(dict_import[key][0], dict_import[key][1], dict_import[key][2], dict_import[key][3],
        #                         dict_import[key][4], dict_import[key][5],
        #                         dict_import[key][6], dict_import[key][7], dict_import[key][8])
    
    return sicro_general_data
    

# comps[key] = Composition(dict_import[key][6], dict_import[key][7])

def get_json_equipments_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_equipments = []
    for key in dict_import:
        for item in dict_import[key][2]:
            equipment = {}
            equipment["comp_code"] = key
            equipment["equipment_code"] = item[0]
            equipment["equipment_quantity"] = item[1]
            equipment["equipment_operative"] = item[2]
            equipment["equipment_inoperative"] = item[3]
            sicro_equipments.append(equipment)
    
    return sicro_equipments

def get_json_labors_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_labors = []
    for key in dict_import:
        for item in dict_import[key][3]:
            labor = {}
            labor["comp_code"] = key
            labor["labor_code"] = item[0]
            labor["labor_quantity"] = item[1]
            sicro_labors.append(labor)
    
    return sicro_labors

def get_json_materials_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_materials = []
    for key in dict_import:
        for item in dict_import[key][4]:
            material = {}
            material["comp_code"] = key
            material["material_code"] = item[0]
            material["material_quantity"] = item[1]
            sicro_materials.append(material)
    
    return sicro_materials

def get_json_auxiliary_activities_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_aux = []
    for key in dict_import:
        for item in dict_import[key][5]:
            aux = {}
            aux["comp_code"] = key
            aux["auxiliary_activity_code"] = item[0]
            aux["auxiliary_activity_quantity"] = item[1]
            sicro_aux.append(aux)
    
    return sicro_aux

def get_json_fixed_time_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_fixed_time = []
    for key in dict_import:
        for item in dict_import[key][6]:
            fixed_time = {}
            fixed_time["comp_code"] = key
            fixed_time["item_code"] = item[0]
            fixed_time["fixed_time_code"] = item[1]
            fixed_time["fixed_time_quantity"] = item[2]
            sicro_fixed_time.append(fixed_time)
    
    return sicro_fixed_time

def get_json_transportation_sicro(filename):

    with open(filename, 'r') as fp:
        dict_import = json.load(fp)
    sicro_transportation = []
    for key in dict_import:
        for item in dict_import[key][7]:
            transportation = {}
            transportation["comp_code"] = key
            transportation["item_code"] = item[0]
            transportation["transportation_quantity"] = item[1]
            transportation["transportation_code_ln"] = item[2]
            transportation["transportation_code_rp"] = item[3]
            transportation["transportation_code_p"] = item[4]
            sicro_transportation.append(transportation)
    
    return sicro_transportation

if __name__=='__main__':
    print(get_json_equipments_sicro('comps.json'))
