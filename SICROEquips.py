import numpy as np
import openpyxl
import tkinter as tk
from tkinter import filedialog
import numpy as np
import pandas as pd

def SOMAComp(CompCode):

    if FIC[CompCode] == None:
        FICValue = 0
    else:
        FICValue = FIC[CompCode]

    CustoUnit = (SomaEquips(CompCode) + SomaMaodeObra(CompCode)) / dict[CompCode][1][0]
    soma = round(CustoUnit + CustoUnit * FICValue + SomaMateriais(CompCode) + SomaAtivAux(CompCode) + SomaTempoFixo(CompCode), 2) # + SomaTransporte(CompCode)

    return soma


def PLANEquipImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps['Sheet1']

    equips = {}

    row = 1

    if sheet['A' + str(row)].value != "Código":
        row = row + 1

    if sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        data7 = sheet['H' + str(row)].value
        data8 = sheet['I' + str(row)].value
        data9 = sheet['J' + str(row)].value
        data10 = sheet['K' + str(row)].value
        equips[code] = (data1, data2, data3, data4, data5, data6, data7, data8, data9, data10)
        row = row + 1

    ps.close()

    return equips

def PLANMaodeObraImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Mão de Obra',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    maodeobra = {}

    row = 1

    if sheet['A' + str(row)].value != "Código":
        row = row + 1

    if sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        data4 = sheet['E' + str(row)].value
        data5 = sheet['F' + str(row)].value
        data6 = sheet['G' + str(row)].value
        maodeobra[code] = (data1, data2, data3, data4, data5, data6)
        row = row + 1

    ps.close()
    return maodeobra

def PLANMaterialImport():

    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Materiais',
                                          filetypes=[("Excel files", "*.xlsx")])

    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    materiais = {}

    row = 1

    if sheet['A' + str(row)].value != "Código":
        row = row + 1

    if sheet['A' + str(row)].value == None:
        row = row + 1

    while row <= sheet.max_row:
        code = sheet['A' + str(row)].value
        data1 = sheet['B' + str(row)].value
        data2 = sheet['C' + str(row)].value
        data3 = sheet['D' + str(row)].value
        materiais[code] = (data1, data2, data3)
        row = row + 1

    ps.close()
    return materiais


def SomaEquips(CompCode):

    soma = 0

    COMPEquips = dict[CompCode][2]
    if len(COMPEquips) != 0:
        for each in COMPEquips:
            EQUIPCode = each[0]
            EQUIPQuant = each[1]
            EQUIPProd = each[2]
            EQUIPImprod = each[3]
            soma = soma + round(EQUIPQuant * (EQUIPProd * equips[EQUIPCode][8] + EQUIPImprod * equips[EQUIPCode][9]), 4)

    return round(soma, 4)

def SomaMaodeObra(CompCode):

    soma = 0
    COMPMaodeObras = dict[CompCode][3]
    if len(COMPMaodeObras) != 0:
        for each in COMPMaodeObras:
            MAODEOBRACode = each[0]
            MAODEOBRAQuant = each[1]
            soma = soma + round(MAODEOBRAQuant * maodeobra[MAODEOBRACode][4], 4)

    return round(soma, 4)

def SomaMateriais(CompCode):

    soma = 0
    COMPMateriais = dict[CompCode][4]
    if len(COMPMateriais) != 0:
        for each in COMPMateriais:
            MATERIALCode = each[0]
            MATERIALQuant = each[1]
            soma = soma + round(MATERIALQuant * materiais[MATERIALCode][2], 4)

    return round(soma, 4)

def SomaAtivAux(CompCode):

    soma = 0
    COMPAtivAux = dict[CompCode][5]
    if len(COMPAtivAux) != 0:
        for each in COMPAtivAux:
            ATIVAUXCode = each[0]
            ATIVAUXQuant = each[1]
            soma = soma + round(ATIVAUXQuant * SOMAComp(ATIVAUXCode), 4)

    return round(soma, 4)

def SomaTempoFixo(CompCode):

    soma = 0
    COMPTempoFixo = dict[CompCode][6]
    if len(COMPTempoFixo) != 0:
        for each in COMPTempoFixo:
            TEMPOFIXOCode = each[1]
            TEMPOFIXOQuant = each[2]
            soma = soma + round(TEMPOFIXOQuant * SOMAComp(TEMPOFIXOCode), 4)

    return round(soma, 4)

def SomaTransporte(CompCode):

    soma = 0
    COMPTransporte = dict[CompCode][7]
    if len(COMPTransporte) != 0:
        for each in COMPTransporte:
            TRANSPORTECode = each[4]
            TRANSPORTEQuant = each[1]
            soma = soma + round(TRANSPORTEQuant * SOMAComp(TRANSPORTECode), 4)

    return round(soma, 4)

equips = PLANEquipImport()
maodeobra = PLANMaodeObraImport()
materiais = PLANMaterialImport()

list = np.load("comps3.npy", allow_pickle=True)
loadFIC = np.load("FIC-PR.npy", allow_pickle=True)
dict = list[()]
FIC = loadFIC[()]

CompCode = '0606785'
print(dict[CompCode][1])
print('Equip', SomaEquips(CompCode))
print('Mão de obra', SomaMaodeObra(CompCode))
print('Material', SomaMateriais(CompCode))
print('Atividades Auxiliares', SomaAtivAux(CompCode))
print('Tempo Fixo', SomaTempoFixo(CompCode))
print('Momento de Transporte', SomaTransporte(CompCode))

soma = SOMAComp(CompCode)
print('Total', soma)
print(soma - SomaTransporte(CompCode))
