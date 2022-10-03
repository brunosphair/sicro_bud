import openpyxl
import tkinter as tk
from tkinter import filedialog

def get_xlsx_materials(filepath=None):
    if not filepath:
        root = tk.Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(title="Selecione o arquivo referente ao Relatório Sintético de Materiais",
                                                filetypes=[("Excel files", "*.xlsx")]) 
    ps = openpyxl.load_workbook(filepath)
    sheet = ps.active

    table_rows = []
    row = 1

    while sheet['A' + str(row)].value != "Código":
        row += 1

    row += 1

    while sheet['A' + str(row)].value is None:
        row += 1

    while row <= sheet.max_row:
        materials = {}
        materials["code"] = sheet['A' + str(row)].value
        materials["description"] = sheet['B' + str(row)].value
        materials["unit"] = sheet['C' + str(row)].value
        if sheet['D' + str(row)].value == '-':
            materials["price"] = 0
        else:
            materials["price"] = sheet['D' + str(row)].value
        table_rows.append(materials)
        row = row + 1

    ps.close()

    return table_rows

def get_xlsx_equipments(filepath=None):
        '''
        Imports the equipments data from the *.xlsx file to a list of dictionaries.
        '''

        if not filepath:
            root = tk.Tk()
            root.withdraw()
            filepath = filedialog.askopenfilename(title='Selecione o arquivo referente ao Relatório Sintético de Equipamentos',
                                                filetypes=[("Excel files", "*.xlsx")])

        ps = openpyxl.load_workbook(filepath)
        sheet = ps.active

        table_rows = []
        row = 1

        while sheet['A' + str(row)].value != "Código":
            row += 1
        row += 1

        while sheet['A' + str(row)].value is None:
            row += 1

        while row <= sheet.max_row:
            equips = {}
            equips["code"] = sheet['A' + str(row)].value
            equips["description"] = sheet['B' + str(row)].value
            equips["acquisition_price"] = sheet['C' + str(row)].value
            equips["depreciation"] = sheet['D' + str(row)].value
            equips["capital_opportunity"] = sheet['E' + str(row)].value
            equips["insurance_and_taxes"] = sheet['F' + str(row)].value
            equips["maintenance"] = sheet['G' + str(row)].value
            equips["operation"] = sheet['H' + str(row)].value
            equips["operation_labor"] = sheet['I' + str(row)].value
            equips["productive_cost"] = sheet['J' + str(row)].value
            equips["unproductive_cost"] = sheet['K' + str(row)].value
            equips[equips["code"]] = (equips["code"],
                                      equips["description"],
                                      equips["acquisition_price"],
                                      equips["depreciation"],
                                      equips["capital_opportunity"],
                                      equips["insurance_and_taxes"],
                                      equips["maintenance"],
                                      equips["operation"],
                                      equips["operation_labor"],
                                      equips["productive_cost"],
                                      equips["unproductive_cost"])
            table_rows.append(equips)
            row = row + 1

        ps.close()

        return table_rows